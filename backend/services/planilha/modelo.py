"""Geracao dos arquivos-modelo que o usuario baixa antes de importar.

Sao .xlsx com o cabecalho exato que o importador espera, mais uma linha de
exemplo. Baixar o modelo e a forma mais simples de o usuario nao errar o
nome das colunas.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font


def _montar(colunas, exemplo, nome_aba):
    planilha = Workbook()
    aba = planilha.active
    aba.title = nome_aba

    aba.append(colunas)
    for celula in aba[1]:
        celula.font = Font(bold=True)

    aba.append(exemplo)

    for indice, coluna in enumerate(colunas, start=1):
        aba.column_dimensions[aba.cell(row=1, column=indice).column_letter].width = (
            max(16, len(coluna) + 6)
        )

    buffer = io.BytesIO()
    planilha.save(buffer)
    buffer.seek(0)
    return buffer


def modelo_alunos():
    """Colunas lidas por ImportarAlunosService."""
    return _montar(
        ["nome", "matricula", "email"],
        ["Maria Silva Santos", "2026001", "maria.santos@escola.com"],
        "Alunos",
    )


def modelo_notas(alunos=None):
    """Colunas lidas por ImportarNotasService.

    Quando a lista de alunos e passada, o modelo ja vem preenchido com eles,
    faltando so a coluna nota - que e o jeito mais rapido de o professor
    lancar a turma inteira.
    """
    planilha = Workbook()
    aba = planilha.active
    aba.title = "Notas"

    aba.append(["matricula", "aluno", "nota", "observacao"])
    for celula in aba[1]:
        celula.font = Font(bold=True)

    if alunos:
        for aluno in alunos:
            aba.append([aluno.get("matricula") or "", aluno.get("nome"), "", ""])
    else:
        aba.append(["2026001", "Maria Silva Santos", "8,5", ""])

    for indice in range(1, 5):
        aba.column_dimensions[aba.cell(row=1, column=indice).column_letter].width = 22

    buffer = io.BytesIO()
    planilha.save(buffer)
    buffer.seek(0)
    return buffer
