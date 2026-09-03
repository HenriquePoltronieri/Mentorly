"""Leitura de planilhas CSV e XLSX enviadas por upload.

Devolve sempre a mesma estrutura, independente do formato: uma lista de
dicts com as chaves em minusculo, mais o numero da linha no arquivo (para a
mensagem de erro apontar exatamente onde o usuario errou).
"""

import csv
import io

from openpyxl import load_workbook

EXTENSOES_ACEITAS = (".csv", ".xlsx")


class PlanilhaInvalida(Exception):
    """Arquivo que nem chega a ser processado linha a linha."""


def _normalizar_cabecalho(valor):
    return str(valor or "").strip().lower()


def _ler_csv(conteudo):
    # Planilha exportada do Excel no Brasil costuma vir em cp1252 e com ';'.
    for codificacao in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            texto = conteudo.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise PlanilhaInvalida("Nao foi possivel ler o arquivo CSV")

    amostra = texto[:2048]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t")
        delimitador = dialeto.delimiter
    except csv.Error:
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","

    leitor = csv.reader(io.StringIO(texto), delimiter=delimitador)
    return [linha for linha in leitor]


def _ler_xlsx(conteudo):
    try:
        planilha = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception:
        raise PlanilhaInvalida("Nao foi possivel abrir a planilha XLSX")

    aba = planilha.active
    return [list(linha) for linha in aba.iter_rows(values_only=True)]


def ler_planilha(nome_arquivo, conteudo):
    """Le o arquivo e devolve [(numero_da_linha, {coluna: valor}), ...].

    A primeira linha nao vazia e tratada como cabecalho.
    """
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".xlsx"):
        linhas = _ler_xlsx(conteudo)
    elif nome.endswith(".csv"):
        linhas = _ler_csv(conteudo)
    else:
        raise PlanilhaInvalida("Envie um arquivo .csv ou .xlsx")

    # Descarta linhas totalmente vazias, mas guarda o numero original para
    # que o erro devolvido aponte a linha certa do arquivo do usuario.
    numeradas = [
        (numero, linha)
        for numero, linha in enumerate(linhas, start=1)
        if any(str(c or "").strip() for c in linha)
    ]
    if not numeradas:
        raise PlanilhaInvalida("A planilha esta vazia")

    _, cabecalho_bruto = numeradas[0]
    colunas = [_normalizar_cabecalho(c) for c in cabecalho_bruto]
    if not any(colunas):
        raise PlanilhaInvalida("A planilha nao tem cabecalho")

    registros = []
    for numero, linha in numeradas[1:]:
        registro = {}
        for indice, coluna in enumerate(colunas):
            if not coluna:
                continue
            valor = linha[indice] if indice < len(linha) else None
            registro[coluna] = "" if valor is None else str(valor).strip()
        registros.append((numero, registro))

    return colunas, registros


def exigir_colunas(colunas, obrigatorias):
    faltando = [c for c in obrigatorias if c not in colunas]
    if faltando:
        raise PlanilhaInvalida(
            "A planilha precisa ter a(s) coluna(s): %s" % ", ".join(faltando)
        )
